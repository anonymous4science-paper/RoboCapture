# main_vague_instruction_generator_v8.1_buffer_path_fix.py

from openai import OpenAI
import json
import base64
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
import time
import random
import pandas as pd
import shutil

# ==============================================================================
# 1. 全局配置 (Global Configuration)
# ==============================================================================
API_SECRET_KEY = 'sk-zk2d2e720fffc0baf63512cff44a0f00c4868a342f738ede'
BASE_URL = "https://api.zhizengzeng.com/v1/"

if not API_SECRET_KEY:
    print("错误：请在代码中填入您的 API_SECRET_KEY。")
    exit()
client = OpenAI(api_key=API_SECRET_KEY, base_url=BASE_URL)

LLM_PARAMS = {
    "max_tokens": 4096,
    "model": "gemini-2.5-pro-preview-06-05",
    #"model": "gpt-4o",
    "temperature": 0.7,
    "top_p": 1,
}

# [修改] 缓冲区配置简化
BUFFER_EXCEL_PATH = "buffer_examples.xlsx"  # 存放示例的Excel文件
BUFFER_IMAGE_DIR = "buffer_images/"  # 存放缓冲区示例对应图片的文件夹
SAMPLES_FROM_BUFFER = 1  # 每次从缓冲区随机抽取的示例数量
STORE_BACK_PROBABILITY = 0.3  # 将新生成结果存回缓冲区的概率
# [新增] 将列名定义为常量，方便统一修改
BUFFER_IMAGE_COLUMN_NAME = "Image Filename"

# ==============================================================================
# 2. 辅助函数 (Helper Functions)
# ==============================================================================
def consult_llm(messages):
    try:
        response = client.chat.completions.create(messages=messages, **LLM_PARAMS)
        return response.choices[0].message.content
    except Exception as e:
        print(f"调用 API 时发生错误: {e}")
        return f"错误：API调用失败。 {e}"


def build_text_message(role, text):
    return {"role": role, "content": text}


def build_multimodal_message(role, text, main_image_path, example_images_paths=[]):
    content = [{"type": "text", "text": text}]

    for img_path in example_images_paths:
        try:
            with open(img_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}})
        except FileNotFoundError:
            print(f"警告：找不到缓冲区示例图片 {img_path}，已跳过。")

    try:
        with open(main_image_path, "rb") as image_file:
            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
            content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}})
    except FileNotFoundError:
        print(f"错误：找不到主图片 {main_image_path}。")
        return None

    return {"role": role, "content": content}


def parse_vague_instruction(output_text):
    # Support "Vague Instruction:" (EN) and "模糊指令:" (ZH)
    m = re.search(r'(?:Vague Instruction|模糊指令)\s*:\s*(.*)', output_text, re.DOTALL | re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return f"Parsing failed: {output_text}"


def parse_decomposition_output(output_text: str):
    # 1) 仅截取最终输出块
    final_block = None
    m = re.search(
        r'\[BEGIN\s+FINAL\s+OUTPUT\s+FORMAT\](.*?)\[(?:END\s+FINAL\s+OUTPUT\s+FORMAT|最终输出格式结束)\]',
        output_text,
        flags=re.IGNORECASE | re.DOTALL
    )
    if m:
        final_block = m.group(1)
    else:
        # 如果没找到最终块，就退化到全文（不推荐，但兜底）
        final_block = output_text

    # 2) 在最终块内提取 Q 和 A（中英都兼容）
    q_match = re.search(r'(?:Clarification Question\(s\)|澄清问题)\s*:\s*(.*)', final_block, flags=re.IGNORECASE)
    a_match = re.search(r'(?:Clarification Answer\(s\)|澄清回答)\s*:\s*(.*)', final_block, flags=re.IGNORECASE)

    def cleanup(text: str) -> str:
        if text is None:
            return "None"
        t = text.strip()
        # 去掉最外层的[]或""（最多两层）
        for _ in range(2):
            if t.startswith('[') and t.endswith(']'):
                t = t[1:-1].strip()
            if (t.startswith('"') and t.endswith('"')) or (t.startswith('“') and t.endswith('”')):
                t = t[1:-1].strip()
        return t

    q_text = cleanup(q_match.group(1) if q_match else "None")
    a_text = cleanup(a_match.group(1) if a_match else "None")

    # 3) 只保留“最终块”的一份问答
    if q_text.lower() in ("none", "") and a_text.lower() in ("none", ""):
        clarification_pair = "No clarification needed"
    else:
        clarification_pair = f"Q: {q_text} A: {a_text}"

    # 4) 提取原子序列（同样只在最终块里找）
    seq_match = re.search(r'(?:Atomic Action Sequence|原子指令序列)\s*:\s*(.*)', final_block, flags=re.IGNORECASE | re.DOTALL)
    sequence_text = seq_match.group(1).strip() if seq_match else "Failed to parse action sequence"

    # 5) 清理可能残留的 END 标记
    sequence_text = re.sub(
        r'\s*\**\s*\[(?:END\s+FINAL\s+OUTPUT\s+FORMAT|最终输出格式结束)\]\s*\**\s*',
        '',
        sequence_text,
        flags=re.IGNORECASE
    ).strip()

    return clarification_pair, sequence_text



def get_next_buffer_filename(buffer_dir, original_extension):
    """
    扫描缓冲区目录，生成下一个序列化的文件名，如 "buffer_image01.png"。
    """
    os.makedirs(buffer_dir, exist_ok=True)
    max_num = 0
    # 正则表达式，用于匹配 "buffer_image" 后跟数字的文件名
    pattern = re.compile(r'buffer_image(\d+)\..+')

    for filename in os.listdir(buffer_dir):
        match = pattern.match(filename)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num

    # 计算下一个编号
    next_num = max_num + 1

    # 格式化为两位数（例如 01, 02, ..., 10）并拼接原始扩展名
    new_filename = f"buffer_image{next_num:02d}{original_extension}"
    return new_filename


def load_buffer_groups(buffer_path, num_groups=SAMPLES_FROM_BUFFER, *, max_rows_per_group=None):
    """
    从缓冲区 Excel 抽取若干“图片分组”：
    - 每个分组对应一张图片（Image Filename）
    - 分组内包含该图片的所有（或至多 max_rows_per_group 条）指令行
    返回: [{"image_filename": str, "rows": [record, ...]} ...]
    """
    try:
        df = pd.read_excel(buffer_path, engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]

        required = ["Image Filename", "Vagueness Type", "Vague Instruction", "Clarification", "Atomic Sequence"]
        for col in required:
            if col not in df.columns:
                print(f"[缓冲区] 警告：缺少列 '{col}'")
                return []

        # 关键：把合并单元格导致的空白向下填充为该块首行的文件名
        df["Image Filename"] = df["Image Filename"].ffill()
        df["Image Filename"] = df["Image Filename"].astype(str).str.strip()
        df = df[~df["Image Filename"].str.lower().isin(["", "nan", "none"])]

        if df.empty:
            print("[缓冲区] 提示：清洗后没有可用样本。")
            return []

        # 按图片文件名分组
        grouped = df.groupby("Image Filename", sort=False)

        # 只保留缓冲区确实存在图片文件的组
        valid_groups = []
        for fname, gdf in grouped:
            img_path = os.path.join(BUFFER_IMAGE_DIR, fname)
            if not os.path.exists(img_path):
                print(f"[缓冲区] 跳过：图片不存在 -> {img_path}")
                continue
            if max_rows_per_group is not None and len(gdf) > max_rows_per_group:
                gdf = gdf.iloc[:max_rows_per_group]
            valid_groups.append({
                "image_filename": fname,
                "rows": gdf.to_dict("records")
            })

        if not valid_groups:
            print("[缓冲区] 警告：没有图片存在于缓冲区目录，返回空列表。")
            return []

        # 抽组（随机但可控数量）
        k = min(num_groups, len(valid_groups))
        sampled_groups = random.sample(valid_groups, k=k)

        # 打印清晰的调试信息
        print(f"[缓冲区] 已抽取 {k} 个图片分组作为 few-shot：")
        for g in sampled_groups:
            print(f"  - {g['image_filename']} | 指令数: {len(g['rows'])}")

        return sampled_groups

    except FileNotFoundError:
        print(f"[缓冲区] 警告：找不到 {buffer_path}")
        return []
    except Exception as e:
        print(f"[缓冲区] 读取异常：{e}")
        return []

def store_group_to_buffer(buffer_path, image_filename, rows, probability):
    """
    rows: [[img, vagueness_type, vague_instruction, clarification, atomic_seq], ...]
    命中概率则一次性把该图片的所有 rows 写回；返回是否写回成功。
    """
    if not rows:
        return False

    # 组级一次概率判定
    if random.random() >= probability:
        print(f" -> [缓冲区] 组级概率未命中，跳过写回：{image_filename}（{len(rows)}条）")
        return False

    try:
        try:
            wb = load_workbook(buffer_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "Buffer Examples"
            headers = ["Image Filename", "Vagueness Type", "Vague Instruction", "Clarification", "Atomic Sequence"]
            ws.append(headers)

        appended = 0
        for row in rows:
            ws.append([image_filename, row[1], row[2], row[3], row[4]])
            appended += 1

        wb.save(buffer_path)
        print(f" -> [缓冲区] 已整体写回 {appended} 条（图片：{image_filename}）")
        return True

    except Exception as e:
        print(f" -> [缓冲区] 组写回失败：{e}")
        return False

# 放在常量区
FAIL_TOKENS_ZH = ["失败", "解析失败", "未能解析出指令序列", "无法生成", "错误", "无效", "空输出"]
FAIL_TOKENS_EN = ["fail", "failed", "error", "invalid", "unable", "cannot"]

def _norm(s: str) -> str:
    s = (s or "").strip()
    # 去掉所有空白，统一小写，便于检索
    return re.sub(r"\s+", "", s, flags=re.S).lower()

def is_generation_ok(vague_instruction: str) -> bool:
    """步骤4生成是否成功"""
    t = _norm(vague_instruction)
    if not t:
        return False
    for k in FAIL_TOKENS_ZH + FAIL_TOKENS_EN:
        if k in t:
            return False
    # 也可加一些正向信号（可选），例如长度至少 4 个汉字/字母
    return len(t) >= 4

def _looks_like_sequence(s: str) -> bool:
    """
    很轻量的结构性校验：是否像一个带编号的序列。
    命中任意一种即可。
    """
    t = s or ""
    return bool(
        re.search(r"\b1\.", t) or     # 1. 抓取...
        re.search(r"①|１．|1、", t) or
        re.search(r"\[\s*1", t)       # [1, 抓取..., 2, ...]
    )

def is_decomposition_ok(clarification_pair: str, atomic_sequence: str) -> bool:
    """步骤5拆解是否成功（看指令序列为主）"""
    seqn = _norm(atomic_sequence)
    if not seqn:
        return False
    for k in FAIL_TOKENS_ZH + FAIL_TOKENS_EN:
        if k in seqn:
            return False
    # 结构性兜底：至少像个步骤序列
    return _looks_like_sequence(atomic_sequence)

# ==============================================================================
# 3. 提示模板 (Prompt Templates) — Translated to English
# ==============================================================================
PROMPT_STEP1_EXTRACT = """
You are a professional scene analyst.

**IMPORTANT: The two-finger gripper in all images is open and is NOT holding any object. Your entire analysis must reflect this fact.** 
* **Real-World Assumption: You MUST treat the scene as a real-life, functional environment. Do NOT use words like "toy," "miniature," "play," or "simulation" in your description.**

From the image below, extract all the objects on the table, at least ten objects. Combine visual content with general world knowledge to describe each object's attributes and characteristics in detail. Treat the scene in the image as a real-life scene; do not describe it as “toys,” “simulation,” or similar.

For every extracted object, you must describe the following attributes:
- **Object name**: The common name of the object.If there are multiple objects of the same type, use attributes like color, location, or other distinct features to give each a unique name (e.g., “red cup,” “blue cup”).
- **Length**: Estimated length (cm).
- **Width**: Estimated width (cm).
- **Height/Thickness**: Estimated height or thickness (cm).
- **Use**: The common use of the object.
- **Interactive Parts**: Describe the object's interactive parts. This can be a discrete component (like a cooktop knob) or an operable property of the object itself (e.g., a faucet can rotate, a lid can be opened). Please specify the method of interaction, its range, and current state. Example: 'The cooktop's knob can be rotated.Example: 'The faucet's spout can be rotated left and right. The faucet's handle can be rotated to control the water flow and is currently in the OFF position.' If none exist, state 'None'.
- **State**: The current state of the object (e.g., open, closed, full, empty, new, old).
- **Color**: The primary color of the object.
- **Location**: Describe the primary area, container, or supporting surface where the object is located. The goal is to understand the object's general zone from the text alone. Examples: "Inside the dish drainer of the sink unit," "In the central area of the tablecloth," "At the bottom of the sink basin."
- **Relative spatial relations**: Describe the object's specific position within its Location (as defined above), relative to its immediate neighbors. You must explicitly state direct physical contact (e.g., 'inside of', 'on top of', 'leaning against'). Examples: "To the left of the red cup, with their sides touching," "Directly above the pink plate, not in contact," "Immediately in front of the faucet."

Strictly follow the format below for each object with no extra explanations:
Object1 Name: [name]
Object1 Length: [estimate] cm
Object1 Width: [estimate] cm
Object1 Height/Thickness: [estimate] cm
Object1 Use: [description]
Object1 Interactive Parts: [description]
Object1 State: [description]
Object1 Color: [description]
Object1 Location: [description]
Object1 Relative Relations: [description]

Object2 Name: [name]
...
"""

PROMPT_STEP2_OPERABILITY = """
You are a motion-planning engineer for an embodied robot equipped with a two-finger gripper. Based on the object descriptions below—and strictly adhering to the robot’s operational constraints—determine whether each object is operable and, if so, how it can be operated. Treat the scene in the image as a real-life scene; do not describe it as “toys,” “simulation,” or similar.

**Robot specs and principles**:
- **Hardware**: Single arm, two-finger gripper, max payload 1 kg, gripping force 20–140 N, stroke width 0–110 mm (~11 cm).
- **Actions**: Can perform grasp, place, push, pull, rotate, press, flip, pour, cut, scoop, shake, sweep, scrub, insert.
- **Principle A (Safety)**: Strictly follow Asimov’s Three Laws—never harm humans. No contact with humans or other living beings.
- **Principle B (Physics)**: Grasped objects must have sufficient rigidity and a fixed shape. Liquids or granular materials must be handled using containers. Large fixed objects (e.g., grill) cannot be moved.**Large fixed objects (e.g., tables, walls, sinks, washbasins) cannot be moved or operated.**
- **Principle C (Restrictions)**: Do not operate sharp items (e.g., knives), or live electrical components (e.g., power outlets).

Analyze the following object list and output strictly in this format with no extra explanations:
Object1 Name: [name]
Object1 Operable: [Yes/No]
Object1 Operation Method: [If operable, describe the specific method, e.g., grasp from the side and place elsewhere; press the top button. If not operable, explain why.]
...
---
**Objects to analyze:**
{object_descriptions}
"""

PROMPT_STEP3_SCENE_AND_INTENT_ANALYSIS = """
...
Describe the entire scene, analyze all potential human needs and intentions in this context, and infer what tasks humans might ask the robot to perform in this scene.
**Crucially, your analysis should consider a variety of tasks that could involve all or most of the different objects identified. Do not just focus on the most obvious interactions. Think about how each object could be uniquely used to create diverse scenarios.**
**Maximize Object Interaction**: Your analysis must consider tasks that involve **as many different objects as possible**, especially those that are not in the center of attention. How can the objects interact with each other in unusual but logical ways?
**Incorporate Diverse Actions**: Brainstorm tasks that would require a variety of robot actions, including **Grasp, Place, Push, Pull, Rotate, Press,Pour, Flip, Cut**. For example, could an object be *pushed* to make space? 
Output a detailed analysis of human intentions in a single coherent paragraph. Do not use a numbered list.
---
**Comprehensive scene information:**
{scene_info}
"""

PROMPT_STEP4_GENERATION_FROM_ANALYSIS = """
You are a creative instruction designer. Your task is to generate one embodied **vague instruction**.

**Learning examples**:
{learning_examples}
---
**Now the new task**

First, review the detailed and objective scene information below as reference.
**Detailed scene information:**
{scene_info}

Next, focus on the integrated analysis of the entire scene and the potential human intentions.
**Integrated scene & intent analysis (main focus):**
{intent_analysis_paragraph}

Finally, apply the specified vagueness type below to generate an instruction that can be executed in the current scene.
**Vagueness type to apply:**
- **Type name**: {vagueness_type_name}
- **Definition**: {vagueness_type_definition}

Based on all the above context and the learning examples, design a reasonable robot task that aligns with the intent analysis and matches the scene facts, is executable in the current scene, and formulate it as a single embodied vague instruction that strictly fits the specified vagueness type.

Creativity and Diversity Constraints:
Novelty: The generated instruction must be novel.
Object Diversity: Make a deliberate effort to utilize a diverse range of objects from the scene. The overall goal is to generate a set of instructions that covers as many different objects as possible.
Action Diversity: The designed task should create opportunities to use actions other than Grasp and Place. Actively consider how to naturally integrate actions such as Push, Pull, Rotate, Press, Flip and Pour into the task.
Contextual Introduction: Introduce reasonable context only when necessary to ensure the intent of the vague instruction can be uniquely and clearly inferred. When a standalone vague instruction would be too broad or illogical, you can use the first clause of a compound sentence to set a scene or premise, which makes the vague part in the second clause understandable. For example, "The plate needs a good scrub; please use the tool for the job,".
**IMPORTANT RULES:**
* **Real-World Assumption: You MUST treat the scene as a real-life, functional environment. Do NOT use words like "toy," "miniature," "play," or "simulation" in your description.**
Strictly follow this format, outputting **only** the vague instruction without any extra explanation:

Vague Instruction: [your generated vague instruction]
"""

VAGUENESS_TYPES = [
    {"name": "Word Choice Error",
     "definition": "The instruction contains an incorrect word choice that must be corrected using visual context to eliminate ambiguity.",
     "examples": [
         "Take the egg out of the cup and place it on the side.",
         "Flip the steak with the knife so it looks better",
         "Cover the pan with the transparent plate to keep it clean."
     ]},

    {"name": "Grammatical Error",
     "definition": "The instruction’s syntax is incomplete or incorrect, but the intent can usually be understood.",
     "examples": [
         "Put steak grill the empty tray.",
         "Cover pan with transparent, keep clean.",
         "Yellow spoon table move to empty tray",
         "Red boxes both, near grill put on the table."
     ]},

    {"name": "Ambiguous Word (Polysemy)",
     "definition": "The instruction contains a verb or noun with multiple meanings that must be disambiguated using visual context. For instance, in “pick it up,” it might mean pick up an object or answer a phone call.",
     "examples": [
         "Please set the steak on the tray.",
         "Now clear the tray on the right side.",
         "Next, draw the spoon closer to the pan.",
         "Take off the left red condiment tray near the grill."
     ]},

    {"name": "Omitted Object",
     "definition": "The instruction specifies an action but omits the target object. The robot can directly infer the unambiguous target object from the visual context and the instruction alone, thus requiring no clarification Q&A. This contrasts with 'Ambiguous Reference.",
     "examples": [
         "Flip it over on the grill.",
         "Cover properly before the steam escapes",
         "Move to the side of the grill so we have more space."
     ]},

    {"name": "Omitted Action",
     "definition": "The instruction specifies an object but omits the action. For example, “I’m thirsty, the cup,” requires inferring the action “hand me the water cup.”",
     "examples": [
         "The grill is getting crowded, and we need more room; the steak on the grill, please.",
         "The steam will escape if we don’t act quickly; the pot lid next to the pan.",
         "Someone asked for a quick snack before we finish setting the table; that pizza slice near the croissant onto the empty tray."
     ]},

    {"name": "Omitted Degree of Action",
     "definition": "The instruction doesn’t specify the extent/degree of the action. For example, “Pull the stool out for me” doesn’t state how far; the degree should be inferred from the intent.",
     "examples": [
         "Shift the steak on the grill so it cooks evenly.",
         "Move the frying pan with the egg closer to the center.",
         "Push the yellow ladle to the side for more space.",
         "Adjust the pizza slice so it fits better on the plate."
     ]},

    {"name": "Common-Sense / Habit Dependence",
     "definition": "Execution must follow human habits and conventions. For example, “Set the table,” which typically implies “fork on the left, knife on the right.",
     "examples": [
         "Place the mouse in a proper position."
     ]},

    {"name": "Relative Spatial Terms",
     "definition": "The instruction uses relative spatial terms such as 'left of', 'right of', 'in front of', or 'between'. Resolving the command requires first identifying a reference object (an 'anchor') to determine the location of the target object. The name of the target object should not appear in the instruction; instead, it must be referred to by a generic term such as 'the item' or 'the thing'. For example, in 'Hand me the thing to the left of the desk lamp,' the 'desk lamp' is the anchor that must be located before 'the thing to the left' can be identified.",
     "examples": [
         "To make space for serving, gently pull the thing to the left of the frying pan toward the edge of the table.",
         "Push the thing between the croissant and the pizza slice closer to the square tray."
     ]},

    {"name": "Relative Object Attributes",
     "definition": "The instruction specifies the target via comparative attributes (size, color, etc.). For example, “Take the slightly larger box.”",
     "examples": [
         "Move the darker tool closer to the frying pan.",
         "Rotate the longer utensil so it faces the center.",
         "Put the wider piece of food on the empty tray."
     ]},

    {"name": "Conditional Trigger",
     "definition": "The instruction uses if-then logic requiring condition checking before action. For example, “If the light is on, turn it off.”",
     "examples": [
         "If there is a steak on the grill, move it into the empty tray.",
         "If an egg is inside the frying pan, cover it with the pot lid.",
         "If the bamboo steamer lid is beside the steamer, put it back on top.",
         "If the ladle is on the tablecloth, move it closer to the frying pan."
     ]},

    {"name": "Ambiguous Reference",
     "definition": "The instruction uses unclear pronouns (e.g., “it”) requiring clarification questions to determine the referent. For example, “Bring it to me.”",
     "examples": [
         "Put that on the tray.",
         "Push that toward the edge.",
         "Cover it before it cools down."
     ]},

    {"name": "High-Level Instruction",
     "definition": "The instruction describes an end goal or abstract task rather than concrete actions. For example, “Prepare the meeting room,” which must be decomposed into specific steps.",
     "examples": [
         "Prepare the grill for serving.",
         "Get the breakfast ready.",
         "Organize the cooking tools."
     ]},

    {"name": "Subjectivity (Human Preference)",
     "definition": "The completion standard depends on personal aesthetics or preferences. For example, “Arrange the decorations to look nicer.”",
     "examples": [
         "Put the tongs and ladle in a neater way.",
         "Adjust the pot lid so it looks properly placed.",
         "Arrange the steak on the tray so it looks more appetizing."
     ]}
]


# [新增] 步骤5的Prompt模板：原子指令拆解
PROMPT_STEP5_DECOMPOSITION = """
You are a top-tier task planner for embodied intelligence. Your job is to decompose a single vague instruction into an atomic action sequence executable by a two-finger gripper robot.

**Learning examples**:
{learning_examples}

---
**Now the new task**

**Background information**:
{scene_info}

**Vague instruction to decompose**:
{vague_instruction}

**Robot action library definitions**:
- **Grasp(Object A)**: Use the gripper to securely grasp Object A. Preconditions: gripper is empty; Object A is graspable; gripper has moved into contact range of Object A. Effect: Object A is grasped; gripper is no longer empty.
- **Place(Object A, Location B)**: Place the held Object A at Location B. Preconditions: Object A is grasped; gripper has reached Location B. Effect: Object A is not grasped; Object A is at Location B; gripper becomes empty.
- **Push(Object A, Start Pos A, Target Pos B)**: Move Object A from Start A to Target B by pushing. Preconditions: gripper has moved into contact range; Object A is at the start position. Effect: Object A is at the target position.
- **Pull(Object A, Start Pos A, Target Pos B)**: Move Object A from Start A to Target B by pulling. Preconditions: Object A is grasped; Object A is at the start position. Effect: Object A is released; Object A is at the target position.
- **Rotate(Object A, Start State A, End State B)**: Rotate an object (e.g., a knob) from the start state to the end state (e.g., ON to OFF). Preconditions: gripper has moved into contact range; Object A is in the start state. Effect: Object A is in the end state.
- **Press(Object A, Start State A, End State B)**: Apply pressure to Object A. Preconditions: gripper has moved into contact range; gripper is empty. Effect: Object A is pressed.
- **Flip(Object A)**: Rotate the gripper wrist 180 degrees to reorient the currently held Object A. Preconditions: Object A is grasped; the current grasp is stable enough for rotation. Effect: Object A remains grasped, but its orientation is flipped by 180 degrees.
- **Pour(ObjectA, into Object B)**: Tilt the gripper wrist and gradually rotate it to change the orientation of the currently held Object A, so that its contents are released into Object B. Preconditions: Object A is grasped securely by the gripper. Object A contains a pourable substance (e.g., liquid or small granular items). Object B is a stable container positioned appropriately below Object A. Effect: The contents of Object A are transferred into Object B. Object A remains grasped after the action, but its content level decreases accordingly.
- **Cut(ObjectA)**: Slice or divide Object A into smaller pieces by applying a controlled downward and forward motion with the knife. Preconditions:A knife is securely grasped by the gripper.Object A (e.g., vegetable, fruit, or other cuttable item) is placed on a stable, immovable surface such as a cutting board.The current grasp on the knife is stable enough for repetitive downward motions. Effect:Object A is cut into smaller pieces. Object A’s geometry and size are altered, but the knife and cutting surface remain unchanged.
- **Scoop(Object A, using Object B)**: Use Object B to collect a portion of Object A. Preconditions: Object B is grasped; Object A is present in a container. Effect: A portion of Object A is now held in Object B.
- **Shake(Object A)**: Move Object A rapidly back and forth to release or scatter its contents. Preconditions: Object A is grasped; Object A contains material that can be released. Effect: A portion of the contents of Object A is released.
- **Sweep(Place A, using Object B)**: Move Object B (e.g., broom, hand brush, mop) repeatedly across Place A to collect or displace dirt and debris. Preconditions: Object B is grasped; Place A contains dirt or debris that can be moved. Effect: Dirt or debris on Place A is gathered, displaced, or removed by Object B.
- **Scrub(Object A, using Object B)**: Move Object B (e.g., cloth, sponge, brush) forcefully back and forth against Object A to remove dirt or stains. Preconditions: Object B is grasped; Object A has surface dirt, stains, or residue that can be removed; optional cleaning agent (e.g., soap) may be applied. Effect: Dirt, stains, or residue on Object A are reduced or removed; Object A becomes cleaner.
- **Insert(Object A, Location B)**: Place Object A into Location B, typically through an opening, slot, or cavity, until Object A is fully or partially contained within Location B. Preconditions: Object A is grasped; Location B has an opening or cavity suitable for insertion. Effect: Object A is securely positioned inside Location B.

**Note on Tool Usage**: For actions like Grasp, Place, and Push, if the robot performs the action using a tool it is currently holding (e.g., using tongs to grasp a steak), you MUST specify the tool at the end of the action. The format is Action(Parameters, using Tool_Name) . If the action is performed directly with the empty gripper, no tool name is needed. Example: [1. Grasp(Tongs), 2. Grasp(Steak, using Tongs)]
**Note**: For actions like “move the basket a bit to the left” or “pull the stool out,” you **must** use Push or Pull with explicit start and target positions.

**Strictly follow this chain-of-thought structure for reasoning and output**:

**1. Envision the target scene**:
Using all background information and the vague instruction, infer the concrete intent and describe a clear end-state of the scene. If multiple interpretations prevent a clear end-state (e.g., ambiguous reference), design up to three reasonable clarification Q&A pairs to resolve the ambiguity, so that the vague instruction can be converted into a single, clear intent.
Important: Only introduce clarification if it is strictly necessary; questions must directly target the ambiguous points in the instruction. There can be at most three Q&A pairs.
- **Output format**:
Clarification Question(s): [If needed, write the question(s) targeting the ambiguity; otherwise write “None”.]
Clarification Answer(s): [If needed, write the assumed answer(s); otherwise write “None”.]
Final Scene Description: [Describe the scene state after the task is completed.]

**2. Determine objects to operate**:
Based on the final scene, infer which objects must be manipulated.
- **Output format**:
Objects to Operate: [Object1, Object2, ...]

**3. Generate atomic instructions**:
Step by step, generate the atomic action sequence based on the final scene and target objects.
- **Output format**:
Initial Action Sequence: [1..., 2....]

**4. Validate the atomic sequence**:
Using the image and general knowledge, check each step from the previous sequence. For each step, state:
- **Validation Step 1**:
    - **Current Scene**: [State of the scene before this step.]
    - **Reason for Action Choice**: [Why this action is chosen.]
    - **Precondition Check**: [Verify each precondition is satisfied.]
    - **Post-Execution Scene**: [How the scene changes after executing this action.]
- **Validation Step 2**:
    - ... (repeat the same process)
If validation fails, you **must** return to Step 3 and regenerate the sequence.

**5. Final output**:
After completing all reasoning and validation, output the confirmed atomic sequence and any clarification Q&A (if present) in the exact format below. This is the only format you must adhere to—do not include any extra explanations or your reasoning process.

**[BEGIN FINAL OUTPUT FORMAT]**
Clarification Question(s): [question(s) or “None”]
Clarification Answer(s): [answer(s) or “None”]
Atomic Action Sequence: [1..., 2...]
**[END FINAL OUTPUT FORMAT]**
"""

# ==============================================================================
# 4. 主执行逻辑 (Main Execution Logic)
# ==============================================================================
if __name__ == '__main__':
    # --- 设置路径 ---
    image_directory = r"D:\PycharmProject\AmbiguousInstructionGenerate\images"
    output_directory = "output"
    if not os.path.exists(output_directory): os.makedirs(output_directory)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel_path = os.path.join(output_directory, f"final_output_{timestamp}.xlsx")

    # --- 获取图片列表 ---
    try:
        image_paths = [os.path.join(image_directory, f) for f in os.listdir(image_directory) if
                       f.endswith(('.png', '.jpg', '.jpeg'))]
        if not image_paths:
            print(f"错误：在 '{image_directory}' 中没有找到图片文件。")
            exit()
    except FileNotFoundError:
        print(f"错误：图片目录 '{image_directory}' 不存在。")
        exit()

    # --- 创建Excel工作簿 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Instruction Generation & Decomposition"
    headers = ["Image", "Vagueness Type", "Generated Vague Instruction", "Clarification Q&A", "Atomic Action Sequence"]
    ws.append(headers)
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 100

    print(f"初始化完成，将处理 {len(image_paths)} 张图片...")

    # --- 遍历每张图片 ---
    for i, image_path in enumerate(image_paths):
        print(f"\n{'=' * 20} 正在处理第 {i + 1}/{len(image_paths)} 张图片: {os.path.basename(image_path)} {'=' * 20}")

        start_row = ws.max_row + 1

        # --- 步骤 1: 提取物体及其特征 ---
        print("步骤 1: 提取物体及其特征...")
        step1_message = build_multimodal_message("user", PROMPT_STEP1_EXTRACT, image_path)
        step1_output = consult_llm([step1_message])
        # [新增] 打印步骤1的模型原始输出
        print("\n=============== [模型输出: 步骤1 - 物体提取] ===============")
        print(step1_output)
        print("====================== [输出结束] ======================\n")
        print(" -> 步骤1完成")

        # --- 步骤 2: 可操作性分析 ---
        print("步骤 2: 可操作性分析...")
        prompt_step2 = PROMPT_STEP2_OPERABILITY.format(object_descriptions=step1_output)
        step2_message = build_text_message("user", prompt_step2)
        step2_output = consult_llm([step2_message])
        # [新增] 打印步骤2的模型原始输出
        print("\n============== [模型输出: 步骤2 - 可操作性分析] ==============")
        print(step2_output)
        print("====================== [输出结束] ======================\n")
        print(" -> 步骤2完成")

        scene_info = f"Object Descriptions:\n{step1_output}\n\nOperability Analysis:\n{step2_output}"

        # --- 步骤 3: 生成综合分析 ---
        print("步骤 3: 生成场景与意图的综合分析...")
        prompt_step3 = PROMPT_STEP3_SCENE_AND_INTENT_ANALYSIS.format(scene_info=scene_info)
        step3_message = build_text_message("user", prompt_step3)
        intent_analysis_paragraph = consult_llm([step3_message])
        # [新增] 打印步骤3的模型原始输出
        print("\n============= [模型输出: 步骤3 - 综合分析] =============")
        print(intent_analysis_paragraph)
        print("====================== [输出结束] ======================\n")
        print(" -> 步骤3完成")

        print("步骤 4 & 5: 开始生成与拆解指令...")

        successful_results_for_this_image = []
        all_steps_successful_for_this_image = True

        buffer_groups = load_buffer_groups(BUFFER_EXCEL_PATH, num_groups=SAMPLES_FROM_BUFFER, max_rows_per_group=None)

        formatted_examples_step4, formatted_examples_step5 = "", ""
        example_image_paths = []

        if buffer_groups:
            for g in buffer_groups:
                fname = g["image_filename"]
                ex_img_path = os.path.join(BUFFER_IMAGE_DIR, fname)
                if os.path.exists(ex_img_path):
                    example_image_paths.append(ex_img_path)

                for idx, row in enumerate(g["rows"], start=1):
                    vt = row.get("Vagueness Type", "")
                    vi = row.get("Vague Instruction", "")
                    cl = row.get("Clarification", "")
                    seq = row.get("Atomic Sequence", "")

                    formatted_examples_step4 += (
                        f"Example:\n"
                        f"Vagueness Type: {vt}\n"
                        f"Vague Instruction: {vi}\n---\n"
                    )
                    formatted_examples_step5 += (
                        f"Example:\n"
                        f"Vague Instruction: {vi}\n"
                        f"Clarification Q&A: {cl}\n"
                        f"Atomic Action Sequence: {seq}\n---\n"
                    )
        else:
            print("[缓冲区] 提示：未抽到任何图片分组，本轮将不使用 few-shot 示例。")

        # === 进入模糊类型循环，复用同一张 few-shot 图片及其全部指令 ===
        for vagueness_type in VAGUENESS_TYPES:
            print(f"  - 应用类型: '{vagueness_type['name']}'")

            # 步骤 4
            prompt_step4 = PROMPT_STEP4_GENERATION_FROM_ANALYSIS.format(
                learning_examples=formatted_examples_step4 if formatted_examples_step4 else "无",
                scene_info=scene_info,
                intent_analysis_paragraph=intent_analysis_paragraph,
                vagueness_type_name=vagueness_type['name'],
                vagueness_type_definition=vagueness_type['definition']
            )
            step4_message = build_multimodal_message("user", prompt_step4, image_path, example_image_paths)
            if step4_message is None:
                print(f"错误：主图片缺失，跳过类型 '{vagueness_type['name']}'")
                continue
            # 步骤 4
            generation_output = consult_llm([step4_message])
            vague_instruction = parse_vague_instruction(generation_output)

            # 先构造并写入一行到 Excel（无论成败都记录）
            clarification_pair, atomic_sequence = "Parsing failed", "Parsing failed"

            ok4 = is_generation_ok(vague_instruction)
            if ok4:
                # 只有步骤4通过，才进行步骤5
                prompt_step5 = PROMPT_STEP5_DECOMPOSITION.format(
                    learning_examples=formatted_examples_step5 if formatted_examples_step5 else "无",
                    scene_info=scene_info,
                    vague_instruction=vague_instruction
                )
                step5_message = build_multimodal_message("user", prompt_step5, image_path, example_image_paths)
                if step5_message is not None:
                    decomposition_output = consult_llm([step5_message])
                    clarification_pair, atomic_sequence = parse_decomposition_output(decomposition_output)

            # 先写 Excel
            row_data = [
                os.path.basename(image_path),
                vagueness_type['name'],
                vague_instruction,
                clarification_pair,
                atomic_sequence
            ]
            ws.append(row_data)

            # 再做统一成功判定
            ok5 = is_decomposition_ok(clarification_pair, atomic_sequence)
            if not (ok4 and ok5):
                all_steps_successful_for_this_image = False
                if not ok4:
                    print(f"[WARN] 生成失败 -> 类型: {vagueness_type['name']} | 文本: {vague_instruction[:80]}")
                if not ok5:
                    print(f"[WARN] 拆解失败 -> 类型: {vagueness_type['name']} | 序列: {atomic_sequence[:80]}")
            else:
                successful_results_for_this_image.append(row_data)

        NUM_TYPES = len(VAGUENESS_TYPES)

        # [这是新代码块]
        # 只有当“全部类型都成功”时，才进入概率写回
        if all_steps_successful_for_this_image and len(successful_results_for_this_image) == NUM_TYPES:

            # --- [核心修改开始] ---

            # 1. 从原始图片路径获取文件扩展名
            _, original_extension = os.path.splitext(image_path)

            # 2. 调用新函数生成序列化的文件名
            new_buffer_filename = get_next_buffer_filename(BUFFER_IMAGE_DIR, original_extension)
            print(f" -> [缓冲区] 生成新的序列化文件名: {new_buffer_filename}")

            # 3. 使用【新文件名】调用写回函数
            did_write = store_group_to_buffer(
                BUFFER_EXCEL_PATH,
                image_filename=new_buffer_filename,  # <--- 使用新名字
                rows=successful_results_for_this_image,
                probability=STORE_BACK_PROBABILITY
            )

            # 4. 如果写回成功，使用【新文件名】复制图片
            if did_write:
                try:
                    destination_path = os.path.join(BUFFER_IMAGE_DIR, new_buffer_filename)
                    shutil.copy(image_path, destination_path)
                    print(f" -> [缓冲区] 图片已成功复制并重命名为: {destination_path}")
                except Exception as e:
                    print(f" -> [缓冲区] 复制图片到缓冲区失败: {e}")

            # --- [核心修改结束] ---

        else:
            print(" -> [缓冲区] 本图未达到“全部成功”，不参与写回判定。")

        # ... (合并单元格与插图逻辑) ...
        end_row = ws.max_row
        if start_row <= end_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            try:
                img = Image(image_path)
                img.width, img.height = 280, (img.height * 280 / img.width)
                ws.add_image(img, f'A{start_row}')
            except Exception as e:
                print(f"错误：无法插入图片 {image_path}。原因: {e}")
                ws[f'A{start_row}'] = os.path.basename(image_path)

        wb.save(output_excel_path)
        print(f" -> 图片 {os.path.basename(image_path)} 的所有指令及拆解已生成并保存。")

    print(f"\n所有处理完成！最终结果已保存在: {output_excel_path}")
